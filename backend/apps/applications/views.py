from rest_framework import generics, status, permissions
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from django_filters import rest_framework as filters
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from .models import Application
from .serializers import (
    ApplicationSubmitSerializer, ApplicationStatusSerializer,
    ApplicationAdminSerializer, StatusUpdateSerializer
)
from apps.accounts.permissions import IsAdminOrAbove, IsModeratorOrAbove
from .services import ApplicationService


class ApplicationFilter(filters.FilterSet):
    status = filters.CharFilter(field_name='status')
    event = filters.NumberFilter(field_name='event_id')
    region = filters.CharFilter(field_name='region', lookup_expr='icontains')
    gender = filters.CharFilter(field_name='gender')
    submitted_after = filters.DateFilter(field_name='submitted_at', lookup_expr='date__gte')
    submitted_before = filters.DateFilter(field_name='submitted_at', lookup_expr='date__lte')

    class Meta:
        model = Application
        fields = ['status', 'event', 'region', 'gender']


class SubmitApplicationView(generics.CreateAPIView):
    queryset = Application.objects.all()
    serializer_class = ApplicationSubmitSerializer
    permission_classes = [permissions.AllowAny]
    parser_classes = [MultiPartParser, FormParser]

    def perform_create(self, serializer):
        application = serializer.save()

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        instance = serializer.instance
        return Response({
            'application_id': instance.application_id,
            'message': "Arizangiz muvaffaqiyatli yuborildi!",
        }, status=status.HTTP_201_CREATED)


@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def track_application(request, application_id):
    try:
        application = Application.objects.get(application_id=application_id.upper())
        serializer = ApplicationStatusSerializer(application)
        return Response(serializer.data)
    except Application.DoesNotExist:
        return Response({'detail': "Ariza topilmadi"}, status=status.HTTP_404_NOT_FOUND)


class AdminApplicationListView(generics.ListAPIView):
    queryset = Application.objects.select_related('event').all()
    serializer_class = ApplicationAdminSerializer
    permission_classes = [IsModeratorOrAbove]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_class = ApplicationFilter
    search_fields = ['application_id', 'full_name', 'email', 'organization', 'presentation_title']
    ordering_fields = ['submitted_at', 'updated_at', 'full_name', 'status']
    ordering = ['-submitted_at']


class AdminApplicationDetailView(generics.RetrieveDestroyAPIView):
    queryset = Application.objects.select_related('event').all()
    serializer_class = ApplicationAdminSerializer
    permission_classes = [IsModeratorOrAbove]


@api_view(['PATCH'])
@permission_classes([IsModeratorOrAbove])
def update_application_status(request, pk):
    try:
        application = Application.objects.get(pk=pk)
    except Application.DoesNotExist:
        return Response({'detail': "Ariza topilmadi"}, status=status.HTTP_404_NOT_FOUND)

    serializer = StatusUpdateSerializer(data=request.data)
    if serializer.is_valid():
        ApplicationService.update_status(
            application,
            serializer.validated_data['status'],
            serializer.validated_data.get('admin_comment', ''),
            actor=getattr(request.user, 'username', 'admin'),
            translations=serializer.validated_data.get('translations')
        )
        return Response({'detail': "Holat muvaffaqiyatli o'zgartirildi", 'status': application.status})
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['DELETE', 'POST'])
@permission_classes([IsModeratorOrAbove])
def bulk_delete_applications(request):
    ids = request.data.get('ids', [])
    if not isinstance(ids, list) or not ids:
        return Response({'detail': "Iltimos, o'chirish uchun arizalarni tanlang."}, status=status.HTTP_400_BAD_REQUEST)
    
    deleted_count, _ = Application.objects.filter(id__in=ids).delete()
    return Response({'detail': f"{deleted_count} ta ariza muvaffaqiyatli o'chirildi."})


@api_view(['POST'])
@permission_classes([IsModeratorOrAbove])
def bulk_status_applications(request):
    ids = request.data.get('ids', [])
    new_status = request.data.get('status')
    admin_comment = request.data.get('comment', '')

    if not isinstance(ids, list) or not ids:
        return Response({'detail': "Iltimos, arizalarni tanlang."}, status=status.HTTP_400_BAD_REQUEST)
    if not new_status:
        return Response({'detail': "Status tanlanmadi."}, status=status.HTTP_400_BAD_REQUEST)

    applications = Application.objects.filter(id__in=ids)
    updated_count = 0
    actor = getattr(request.user, 'username', 'admin')

    for app in applications:
        ApplicationService.update_status(
            app,
            new_status,
            admin_comment,
            actor=actor
        )
        updated_count += 1

    return Response({'detail': f"{updated_count} ta arizaning holati muvaffaqiyatli o'zgartirildi."})


@api_view(['GET'])
@permission_classes([IsAdminOrAbove])
def export_applications_excel(request):
    apps = Application.objects.select_related('event').filter(
        **{k: v for k, v in request.query_params.items() if k in ['status', 'event_id']}
    ).order_by('-submitted_at')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Arizalar"

    headers = ['Ariza ID', 'Ism sharif', 'Email', 'Telefon', 'Tashkilot', 'Lavozim', 'Viloyat', 'Tadbir', 'Taqdimot mavzusi', 'Holat', 'Yuborilgan']
    header_fill = PatternFill(start_color='1a56db', end_color='1a56db', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    STATUS_LABELS = dict(Application.Status.choices)
    for row, app in enumerate(apps, 2):
        ws.append([
            app.application_id, app.full_name, app.email, app.phone,
            app.organization, app.position, app.region, app.event.title,
            app.presentation_title, STATUS_LABELS.get(app.status, app.status),
            app.submitted_at.strftime('%Y-%m-%d %H:%M'),
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="arizalar.xlsx"'
    wb.save(response)
    return response
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework import permissions
from .models import Application
from apps.accounts.permissions import IsModeratorOrAbove

@api_view(['POST'])
@permission_classes([IsModeratorOrAbove])
def check_in_application(request, pk):
    try:
        app = Application.objects.get(pk=pk)
        app.attended = True
        app.save()
        return Response({'status': 'success', 'message': 'Application checked in successfully.'})
    except Application.DoesNotExist:
        return Response({'status': 'error', 'message': 'Application not found.'}, status=404)
